import json
from enum import unique

from django.contrib import messages
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *

from openpyxl import load_workbook
from django import forms
from io import BytesIO
from django.http import HttpResponse
import openpyxl
from functools import wraps


def superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_superuser:
            messages.warning(request, "You do not have permission to access this page.")
            return redirect('admin_login')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def staff_or_superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            messages.warning(request, "You do not have permission to access this page.")
            return redirect('admin_login')
        return view_func(request, *args, **kwargs)
    return _wrapped_view
def home(request):
        return render(request, 'home.html', {})

def get_registered_queryset_from_filter(filter_field, filter_operator, filter_value):
    try:
        filter_kwargs = {}

        # Boolean-type fields (including breakout as null check)
        if filter_operator.lower() in ["true", "false", "yes", "no"]:
            truthy = filter_operator.lower() in ["true", "yes", "1"]

            # Special handling for ForeignKey fields like 'breakout'
            if filter_field == "breakout":
                filter_kwargs = {f"{filter_field}__isnull": not truthy}
            else:
                filter_kwargs = {f"{filter_field}": truthy}

        # 'ne' means not equal (exclude)
        elif filter_operator == "ne":
            if not filter_value:
                return None, "Filter value is required for 'ne' operator."
            filter_kwargs = {f"{filter_field}__exact": filter_value}
            queryset = Registered.objects.exclude(**filter_kwargs)
            return queryset, None

        # Other operators (icontains, gte, lte, etc.)
        else:
            if not filter_value:
                return None, "Filter value is required for the selected filter."
            filter_kwargs = {f"{filter_field}__{filter_operator}": filter_value}

        queryset = Registered.objects.filter(**filter_kwargs)
        return queryset, None

    except Exception as e:
        return None, str(e)


@superuser_required
def registered(request):
    registered_list = Registered.objects.all().order_by("id")

    # Statistics
    total_count = registered_list.count()
    tag_downloaded_count = registered_list.filter(has_downloaded_tag=True).count()
    confirmed_attendance_count = registered_list.filter(has_confirmed_attendance=True).count()
    present_count = registered_list.filter(is_present=True).count()
    present_count_day2 = registered_list.filter(is_present_day2=True).count()
    breakout_selected_count = registered_list.filter(breakout__isnull=False).count()

    context = {
        'registered_list': registered_list,
        'searched': False,
        'search': '',
        'total_count': total_count,
        'tag_downloaded_count': tag_downloaded_count,
        'confirmed_attendance_count': confirmed_attendance_count,
        'present_count': present_count,
        'present_count_day2': present_count_day2,
        'breakout_selected_count': breakout_selected_count,
    }

    if request.method == "POST":
        field = request.POST.get('filter_field')
        operator = request.POST.get('filter_operator')
        value = request.POST.get('filter_value')

        queryset, error = get_registered_queryset_from_filter(field, operator, value)
        if queryset is not None:
            context['registered_list'] = queryset.order_by("id")
            context['searched'] = True
            context['field'] = field
            context['operator'] = operator
            context['value'] = value
            context['search'] = f"{field} {operator} '{value}'" if value else f"{field} = {operator}"
            context['search_count'] = queryset.count()
        else:
            context['registered_list'] = []
            context['searched'] = True
            context['search'] = error or "Invalid search"

    return render(request, 'viewregistered.html', context)


def register(request):
    if request.user.is_authenticated and request.user.is_superuser:
        if request.method == 'POST':
            form = RegisterForm(request.POST)
            if form.is_valid():
                registered_user = form.save()
                return redirect('thank_you', registered_user.unique_id)
        else:
            form = RegisterForm()
        return render(request, 'register.html', {'form': form,})
    return render(request, 'registration_closed.html', {})

def thank_you(request, unique_id):
    couple = Registered.objects.get(unique_id=unique_id)
    return render(request, 'thank_you.html', {'couple': couple})

@superuser_required
def mark_present(request, unique_id):
    reg_couple = Registered.objects.get(unique_id=unique_id)
    reg_couple.is_present_day2 = True
    reg_couple.save()
    return render(request, 'mark_present.html', {'registered': reg_couple})

@staff_or_superuser_required
def update_registee(request, unique_id):
    reg_couple = Registered.objects.get(unique_id=unique_id)
    form = RegisterForm(request.POST or None, instance=reg_couple)
    if form.is_valid():
        form.save()
        messages.success(request, f"{reg_couple.id} - {reg_couple.s_name} record has been updated successfully!")
        return redirect('view-registered')
    return render(request, 'update_registee.html', {'registered': reg_couple, 'form': form})

@staff_or_superuser_required
def delete_registee(request, unique_id):
    reg_couple = Registered.objects.get(unique_id=unique_id)
    messages.warning(request, f"Record {reg_couple.id} - {reg_couple.s_name} has been deleted.")
    reg_couple.delete()
    return redirect('view-registered')

def pdf_registee(request, unique_id):
    import qrcode
    from PIL import Image
    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader
    from django.http import FileResponse, HttpResponse
    from django.contrib import messages
    from django.shortcuts import redirect

    reg_couple = Registered.objects.get(unique_id=unique_id)

    # Check if breakout is selected
    if not reg_couple.breakout:
        messages.warning(request, "You must register for a breakout before downloading your tag.")
        return redirect('choose_breakout', unique_id=unique_id)

    # Check if attendance has been confirmed
    if not reg_couple.has_confirmed_attendance:
        messages.warning(request, "You must confirm attendance before downloading your tag.")
        return redirect('confirm_attendance', surname=reg_couple.s_name, unique_id=unique_id)

    # Page setup
    page_width, page_height = A4
    tag_width = 8.8 * cm
    tag_height = 12.4 * cm
    x_left = (page_width - (2 * tag_width + 1 * cm)) / 2
    x_right = x_left + tag_width + 1 * cm
    top_margin = 0.5 * cm
    y_position = page_height - tag_height - top_margin

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    # Background image setup
    image_path = 'static/tag_background.png'
    bg_image = Image.open(image_path).convert("RGB")
    bg_resized = bg_image.resize((int(tag_width), int(tag_height)))
    bg_io = io.BytesIO()
    bg_resized.save(bg_io, format='PNG')
    bg_io.seek(0)
    bg_reader = ImageReader(bg_io)

    # QR code generation
    qr_data = request.build_absolute_uri(f"/mark_present/{reg_couple.unique_id}")
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white')
    qr_io = io.BytesIO()
    qr_img.save(qr_io, format='PNG')
    qr_io.seek(0)
    qr_reader = ImageReader(qr_io)

    # Positions for each tag
    positions = [
        {"x": x_left, "y": y_position, "name": reg_couple.f_name_f.upper()},
        {"x": x_right, "y": y_position, "name": reg_couple.f_name_m.upper()}
    ]

    for pos in positions:
        # Draw tag background
        c.drawImage(bg_reader, pos["x"], pos["y"], width=tag_width, height=tag_height)

        # Draw border
        c.setStrokeColorRGB(0, 0, 0)
        c.rect(pos["x"], pos["y"], tag_width, tag_height)

        center_x = pos["x"] + tag_width / 2
        content_start_y = pos["y"] + 4.8 * cm

        # Surname
        c.setFont("Times-Bold", 18)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(center_x, content_start_y, reg_couple.s_name.lower())

        # First name
        c.drawCentredString(center_x, content_start_y - 0.6 * cm, pos["name"])

        # QR code
        qr_size = 2.8 * cm
        qr_x = center_x - qr_size / 2
        qr_y = content_start_y - 3.8 * cm
        c.drawImage(qr_reader, qr_x, qr_y, width=qr_size, height=qr_size)

        # Breakout code name under QR (in red and bold)
        breakout_code = reg_couple.breakout.code_name
        c.setFont("Times-Bold", 12)
        c.setFillColorRGB(1, 0, 0)  # Red
        c.drawCentredString(center_x, qr_y - 0.55 * cm, f"{breakout_code}")

    c.showPage()
    c.save()
    buf.seek(0)

    reg_couple.has_downloaded_tag = True
    reg_couple.save()

    return FileResponse(buf, as_attachment=True, filename=f'{reg_couple.s_name}.pdf')


def download_tag(request, surname, unique_id):
    reg_couple = get_object_or_404(Registered, s_name__iexact=surname, unique_id=unique_id)
    return render(request, 'download_tag.html', {'registered': reg_couple})

@staff_or_superuser_required
def view_registee(request, unique_id):
    reg_couple = get_object_or_404(Registered, unique_id=unique_id)
    reg_code = generate_stable_code(unique_id)
    return render(request, 'view_registee.html', {'registered': reg_couple, 'reg_code': reg_code})

def confirm_attendance(request, surname, unique_id):
    reg_couple = get_object_or_404(Registered, s_name__iexact=surname, unique_id=unique_id)

    # Handle POST request
    if request.method == 'POST':
        if request.GET.get('action') == 'unconfirm':
            if request.user.is_authenticated and request.user.is_superuser:
                reg_couple.has_confirmed_attendance = False
                reg_couple.breakout = None
                reg_couple.save()
                messages.success(request, "Attendance Unconfirmed and Breakout Selection Removed Successfully")
            else:
                messages.error(request, "You are not authorized to perform this action.")
        else:
            reg_couple.has_confirmed_attendance = True
            reg_couple.save()
        return redirect('confirm_attendance', surname=surname, unique_id=unique_id)

    return render(request, 'confirm_attendance.html', {'registered': reg_couple})

@superuser_required
def view_comments(request):
    couples_with_comments = Registered.objects.filter(comments__isnull=False).exclude(comments__exact='').order_by('-id')
    return render(request, 'view_comments.html', {'couples': couples_with_comments})

def resources(request):
    resource_info = Resource.objects.all().order_by('-priority', '-id')
    return render(request, 'resources/resource_list.html', {'resources': resource_info})

@superuser_required
def add_resource(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        link = request.POST.get('link')
        priority = 'priority' in request.POST

        Resource.objects.create(name=name, description=description, link=link, priority=priority)
        return redirect('manage_resources')

    return render(request, 'resources/add_resource.html')

@superuser_required
def manage_resources(request):
    resource_info = Resource.objects.all().order_by('-priority', '-id')
    return render(request, 'resources/manage_resources.html', {'resources': resource_info})

@superuser_required
def edit_resource(request, pk):
    resource_info = get_object_or_404(Resource, pk=pk)
    if request.method == 'POST':
        resource_info.name = request.POST.get('name')
        resource_info.description = request.POST.get('description')
        resource_info.link = request.POST.get('link')
        resource_info.priority = 'priority' in request.POST
        resource_info.save()
        messages.success(request, f'Resource ({resource_info.name}) has been updated.')
        return redirect('manage_resources')

    return render(request, 'resources/edit_resource.html', {'resource': resource_info})

@superuser_required
def delete_resource(request, pk):
    resource_info = get_object_or_404(Resource, pk=pk)
    resource_name = resource_info.name
    resource_info.delete()
    messages.success(request, f"Resource {resource_name} deleted successfully.")
    return redirect('manage_resources')

class ExcelUploadForm(forms.Form):
    file = forms.FileField(label="Upload Excel File (.xlsx only)")

@superuser_required
def upload_excel(request):
    success_count = 0
    error_rows = []
    required_fields = [
        's_name', 'f_name_m', 'phone_no_m', 'email_m',
        'f_name_f', 'phone_no_f', 'email_f', 'year_married',
        'attended_previous', 'how_heard_about_program', 'comments'
    ]

    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']

            try:
                wb = load_workbook(filename=BytesIO(excel_file.read()))
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]

                if not all(field in headers for field in required_fields):
                    messages.error(request, "The uploaded file is missing one or more required columns.")
                    return redirect('upload_excel')

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    data = dict(zip(headers, row))

                    def normalize_name(name):
                        return name.strip().title().replace(' ', '-') if name else ''

                    def normalize_phone(number):
                        if not number:
                            return ""
                        number = str(number).strip()

                        # Handle local 11-digit Nigerian numbers like 08012345678
                        if number.isdigit() and len(number) == 11 and number.startswith(("07", "08", "09")):
                            number = "+234" + number[1:]

                        # Handle numbers starting with 7, 8, 9 (likely missing 0 or country code)
                        elif number.isdigit() and len(number) == 10 and number.startswith(("7", "8", "9")):
                            number = "+234" + number

                        elif number.startswith("234") and len(number) >= 13:
                            number = "+" + number

                        elif not number.startswith("+"):
                            number = "+" + number

                        return number

                    def normalize_choice(value, valid_choices):
                        value = str(value).strip().title()
                        return value if value in valid_choices else 'Other'

                    def normalize_attended(value):
                        value = str(value).strip().title()
                        return "Yes" if value == "Yes" else "No"

                    cleaned = {
                        's_name': normalize_name(data.get('s_name')),
                        'f_name_m': normalize_name(data.get('f_name_m')),
                        'phone_no_m': normalize_phone(data.get('phone_no_m')),
                        'email_m': data.get('email_m', '').strip().lower() if data.get('email_m') else '',
                        'f_name_f': normalize_name(data.get('f_name_f')),
                        'phone_no_f': normalize_phone(data.get('phone_no_f')),
                        'email_f': data.get('email_f', '').strip().lower() if data.get('email_f') else '',
                        'year_married': data.get('year_married'),
                        'attended_previous': normalize_attended(data.get('attended_previous')),
                        'how_heard_about_program': normalize_choice(data.get('how_heard_about_program'), ['Flyer', 'Friend', 'Church', 'Social Media', 'Other']),
                        'comments': data.get('comments', '')
                    }

                    form = RegisterForm(cleaned)
                    if form.is_valid():
                        try:
                            form.save()
                            success_count += 1
                        except Exception as save_err:
                            error_rows.append((row_idx, f"DB save error: {str(save_err)}"))
                    else:
                        error_rows.append((row_idx, form.errors.as_text()))

                messages.success(request, f"{success_count} records imported successfully.")

            except Exception as e:
                messages.error(request, f"Failed to process file: {str(e)}")

    else:
        form = ExcelUploadForm()

    return render(request, "upload_excel.html", {
        'form': form,
        'error_rows': error_rows,
        'success_count': success_count,
        'required_fields': required_fields
    })

@superuser_required
def export_registered_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Couples"

    headers = [
        "Unique ID", "Registration Code", "Surname", "Husband's First Name", "Phone No (H)", "Email (H)",
        "Wife's First Name", "Phone No (W)", "Email (W)", "Year Married",
        "Attended Before?", "Heard About Program", "Comments",
        "Labourer", "Downloaded Tag?", "Confirmed Attendance?", "Present?", "Present 2?", "Breakout Attended"
    ]
    ws.append(headers)

    for couple in Registered.objects.all():
        code = generate_stable_code(couple.unique_id)
        row = [
            str(couple.unique_id),
            code,
            couple.s_name,
            couple.f_name_m,
            couple.phone_no_m,
            couple.email_m if couple.email_m else "",
            couple.f_name_f,
            couple.phone_no_f,
            couple.email_f if couple.email_f else "",
            couple.year_married,
            couple.attended_previous,
            couple.how_heard_about_program,
            couple.comments if couple.comments else "",
            couple.labourer.username if couple.labourer else "",
            "Yes" if couple.has_downloaded_tag else "No",
            "Yes" if couple.has_confirmed_attendance else "No",
            "Yes" if couple.is_present else "No",
            "Yes" if couple.is_present_day2 else "No",
            couple.breakout.title if couple.breakout else "",
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registered_couples.xlsx'
    wb.save(response)
    return response

def privacy_policy(request):
    return render(request, "privacy_policy.html", {})

@superuser_required
def breakout_admin_dashboard(request):
    breakouts = Breakouts.objects.all().order_by('title')
    for breakout in breakouts:
        breakout.assigned_couples = breakout.registered_set.all()
        breakout.assigned_count = breakout.assigned_couples.count()

    unassigned = Registered.objects.filter(breakout__isnull=True).order_by('s_name')

    return render(request, 'breakouts/admin_dashboard.html', {
        'breakouts': breakouts,
        'unassigned': unassigned,
        'unassigned_count': unassigned.count()
    })

@superuser_required
def create_breakout(request):
    if request.method == 'POST':
        form = BreakoutForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Breakout created successfully.")
            return redirect('breakout_admin_dashboard')
    else:
        form = BreakoutForm()
    return render(request, 'breakouts/breakout_form.html', {'form': form, 'title': 'Create Breakout'})

@superuser_required
def edit_breakout(request, breakout_id):
    breakout = get_object_or_404(Breakouts, id=breakout_id)
    if request.method == 'POST':
        form = BreakoutForm(request.POST, instance=breakout)
        if form.is_valid():
            form.save()
            messages.success(request, "Breakout updated successfully.")
            return redirect('breakout_admin_dashboard')
    else:
        form = BreakoutForm(instance=breakout)
    return render(request, 'breakouts/breakout_form.html', {'form': form, 'title': 'Edit Breakout'})

@superuser_required
def delete_breakout(request, breakout_id):
    breakout = get_object_or_404(Breakouts, id=breakout_id)
    assigned_count = breakout.registered_set.count()

    if request.method == 'POST':
        breakout.delete()
        messages.success(request, f"Breakout deleted. {assigned_count} couple(s) were unassigned.")
        return redirect('breakout_admin_dashboard')

    return render(request, 'breakouts/confirm_delete.html', {
        'breakout': breakout,
        'assigned_count': assigned_count
    })

def choose_breakout(request, unique_id):
    couple = get_object_or_404(Registered, unique_id=unique_id)
    user = request.user
    breakouts = Breakouts.objects.all().order_by('title')

    is_locked = couple.breakout is not None and not (user.is_superuser or user.is_staff)

    if request.method == 'POST':
        selected_id = request.POST.get('breakout_id')
        selected_breakout = Breakouts.objects.get(id=selected_id)

        if is_locked:
            return redirect('choose_breakout', unique_id=unique_id)

        if selected_breakout.slots_remaining() <= 0:
            messages.error(request, f"'{selected_breakout.title}' is already full. Please select another breakout.")
            return redirect('choose_breakout', unique_id=unique_id)

        couple.breakout = selected_breakout
        couple.save()
        return redirect('choose_breakout', unique_id=unique_id)

    return render(request, 'breakouts/choose_breakout.html', {
        'couple': couple,
        'breakouts': breakouts,
        'is_locked': is_locked,
    })

def couple_welcome(request, surname, unique_id):
    try:
        couple = Registered.objects.get(unique_id=unique_id, s_name__iexact=surname)
    except Registered.DoesNotExist:
        messages.warning(request, "No matching registration found. Please verify your link or contact us.")
        return redirect('home')

    return render(request, 'couple_welcome.html', {'couple': couple})

def auth_lookup(request):
    context = {}
    if request.method == "POST":
        surname_input = request.POST.get("surname", "")
        phone_input = request.POST.get("phone", "").strip()

        normalized_surname = surname_input.strip().replace(" ", "-").lower()
        from django.db.models.functions import Lower

        couple = Registered.objects.filter(
            Q(phone_no_m=phone_input) | Q(phone_no_f=phone_input)
        ).annotate(
            normalized_db_surname=Lower('s_name')
        ).filter(
            normalized_db_surname=normalized_surname
        ).first()

        # Save session and redirect to code entry
        if couple:
            request.session["auth_uuid"] = str(couple.unique_id)
            request.session["auth_phone"] = phone_input
            return redirect("auth_code")
        else:
            context = {'error': "No matching record found. Please check your surname and phone number.",
                       'surname': surname_input,
                       'phone': phone_input,}

    return render(request, "auth_lookup.html", context)

def auth_code(request):
    couple_uuid = request.session.get("auth_uuid")
    if not couple_uuid:
        messages.error(request, "Session expired. Please try again.")
        return redirect("couple_lookup")

    try:
        couple = Registered.objects.get(unique_id=couple_uuid)
    except Registered.DoesNotExist:
        messages.error(request, "Invalid session data.")
        return redirect("couple_lookup")

    if request.method == "POST":
        code_entered = request.POST.get("code").strip().upper()
        correct_code = generate_stable_code(couple.unique_id)

        if code_entered == correct_code:
            return redirect("couple_welcome", surname=couple.s_name,unique_id=couple.unique_id)
        else:
            messages.error(request, "Invalid code. Please try again.")

    return render(request, "auth_code.html", {"couple": couple})

import hashlib
def generate_stable_code(uuid_val):
    """
    Generates a deterministic 6-character code from a UUID.
    The same UUID will always return the same code.
    """
    h = hashlib.sha256(str(uuid_val).encode()).hexdigest()
    return h[:6].upper()  # Return the first 6 hex digits in uppercase

from django.http import JsonResponse
def send_auth_code(request):
    couple_uuid = request.session.get("auth_uuid")
    phone_used = request.session.get("auth_phone")

    if not couple_uuid and not phone_used:
        return JsonResponse({"status": "error", "message": "Session expired. Please go back and try again."})

    try:
        couple = Registered.objects.get(unique_id=couple_uuid)
    except Registered.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Invalid session data."})

    result = send_verification_code(couple, phone_used)
    return JsonResponse(result)

from messaging_engine.messaging import send_verification_email, send_verification_sms
def send_verification_code(couple, input_phone):
    code = generate_stable_code(couple.unique_id)

    # Determine which contact to send to
    if couple.phone_no_m == input_phone:
        if couple.email_m:
            success, info = send_verification_email(couple.s_name, couple.f_name_m, couple.f_name_f, couple.email_m, code)
            return {"status": "success" if success else "error", "message": info}
        elif couple.phone_no_m:
            success, info = send_verification_sms(couple.s_name, couple.f_name_m, couple.phone_no_m, code)
            return {"status": "success" if success else "error", "message": info}

    elif couple.phone_no_f == input_phone:
        if couple.email_f:
            success, info = send_verification_email(couple.s_name, couple.f_name_m, couple.f_name_f, couple.email_f, code)
            return {"status": "success" if success else "error", "message": info}
        elif couple.phone_no_f:
            success, info = send_verification_sms(couple.s_name, couple.f_name_f, couple.phone_no_f, code)
            return {"status": "success" if success else "error", "message": info}

    return {"status": "error", "message": "No valid contact info found for the provided phone number."}


def submit_question(request, surname, unique_id):
    remembered_questions = Question.objects.filter(
        surname__iexact=surname,
        unique_id=unique_id,
        remember=True
    ).order_by('created_at')

    if request.method == 'POST':
        text = request.POST.get('text', '')
        remember = 'remember' in request.POST

        Question.objects.create(
            surname=surname,
            unique_id=unique_id,
            text=text,
            remember=remember
        )
        messages.success(request, "Your question was submitted.")
        return redirect('submit_question', surname=surname, unique_id=unique_id)

    return render(request, 'questions/submit_question.html', {
        'questions': remembered_questions,
        'surname': surname,
        'unique_id': unique_id
    })

def edit_question(request, pk, surname, unique_id):
    if request.user.is_authenticated and request.user.is_superuser:
        question = get_object_or_404(Question, pk=pk, surname__iexact=surname, unique_id=unique_id)
    else:
        question = get_object_or_404(Question, pk=pk, surname__iexact=surname, unique_id=unique_id, remember=True)

    if request.method == 'POST':
        question.text = request.POST.get('text', question.text)
        question.remember = 'remember' in request.POST
        question.save()
        messages.success(request, "Your question has been updated.")
        if request.user.is_authenticated and request.user.is_superuser:
            return redirect('admin_question_list')
        return redirect('submit_question', surname=surname, unique_id=unique_id)

    return render(request, 'questions/edit_question.html', {
        'question': question,
        'surname': surname,
        'unique_id': unique_id,
    })

def delete_question(request, pk, surname, unique_id):
    question = get_object_or_404(Question, pk=pk, surname__iexact=surname, unique_id=unique_id)
    question.delete()
    messages.success(request, "Your question was deleted.")
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('admin_question_list')
    return redirect('submit_question', surname=surname, unique_id=unique_id)

@superuser_required
def list_all_questions(request):
    questions = Question.objects.all().order_by('created_at')
    return render(request, 'questions/admin_question_list.html', {'questions': questions})

@superuser_required
def present_questions(request):
    from django.core.paginator import Paginator
    questions = Question.objects.all().order_by('created_at')
    q = request.GET.get('q', '1')

    try:
        q = int(q)
        if q < 1:
            raise ValueError
    except ValueError:
        return redirect(f"{request.path}?q=1")

    total = questions.count()

    if total == 0:
        return render(request, 'questions/present_questions.html', {'no_questions': True})

    if q > total:
        return redirect(f"{request.path}?q={total}")

    question = questions[q - 1]
    context = {
        'question': question,
        'q_index': q,
        'total': total,
        'first_q': 1,
        'last_q': total,
        'has_prev': q > 1,
        'has_next': q < total,
    }

    return render(request, 'questions/present_questions.html', context)

def submit_feedback(request, surname, unique_id):
    reg = get_object_or_404(Registered, unique_id=unique_id, s_name__iexact=surname)

    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.unique_id = reg.unique_id
            feedback.surname = reg.s_name
            feedback.registration_code = generate_stable_code(reg.unique_id)
            feedback.save()
            return redirect('feedback_thank_you', reg.unique_id)  # make sure to create this
    else:
        form = FeedbackForm()

    return render(request, 'feedback_form.html', {
        'form': form,
        'couple': reg,
    })

def feedback_thank_you(request, unique_id):
    couple = Registered.objects.get(unique_id=unique_id)
    return render(request, 'feedback_thank_you.html', {'couple': couple})

@superuser_required
def feedback_list(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')
    return render(request, 'feedback_list.html', {'feedbacks': feedbacks})

@superuser_required
def export_feedback_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Responses"

    # Header row
    headers = [
        "Surname", "Registration Code", "Unique ID",
        "Satisfaction", "Website Usability", "Completion Experience",
        "Suggestions", "Seminar Ideas", "Submitted At"
    ]
    ws.append(headers)

    # Populate rows
    for fb in Feedback.objects.all().order_by('-submitted_at'):
        row = [
            fb.surname,
            fb.registration_code,
            str(fb.unique_id),
            fb.satisfaction,
            fb.usability,
            fb.completion,
            fb.suggestions or "",
            fb.seminar_ideas or "",
            fb.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
        ]
        ws.append(row)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=retreat_feedback.xlsx'
    wb.save(response)
    return response